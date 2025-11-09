import pyautogui
import time
import pyperclip
import google.generativeai as genai
import webbrowser
import json
import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
import os


def ler_temas_existentes(arquivo_planilha):
    """Lê todos os temas já existentes na planilha para evitar duplicatas."""
    temas_existentes = set()
    
    if not os.path.exists(arquivo_planilha):
        return temas_existentes
    
    try:
        workbook = load_workbook(arquivo_planilha)
        worksheet = workbook.active
        
        # Pula a linha de cabeçalho (linha 1)
        for row in range(2, worksheet.max_row + 1):
            tema = worksheet.cell(row, 1).value
            if tema and isinstance(tema, str):
                # Normaliza o tema para comparação (minúsculas, remove espaços extras)
                tema_normalizado = tema.lower().strip()
                temas_existentes.add(tema_normalizado)
        
        return temas_existentes
    except Exception as e:
        print(f"⚠️ Erro ao ler temas existentes: {e}")
        return temas_existentes


def temas_sao_similares(tema1, tema2):
    """Verifica se dois temas são muito similares (evita duplicatas com pequenas variações)."""
    t1 = tema1.lower().strip()
    t2 = tema2.lower().strip()
    
    # Se forem idênticos após normalização
    if t1 == t2:
        return True
    
    # Se um contém o outro (com diferença mínima)
    palavras_t1 = set(t1.split())
    palavras_t2 = set(t2.split())
    
    # Se compartilham mais de 70% das palavras principais (palavras com mais de 3 caracteres)
    palavras_principais_t1 = {p for p in palavras_t1 if len(p) > 3}
    palavras_principais_t2 = {p for p in palavras_t2 if len(p) > 3}
    
    if palavras_principais_t1 and palavras_principais_t2:
        palavras_comuns = palavras_principais_t1 & palavras_principais_t2
        todas_palavras = palavras_principais_t1 | palavras_principais_t2
        if todas_palavras and len(palavras_comuns) / len(todas_palavras) > 0.7:
            return True
    
    return False


def filtrar_temas_repetidos(temas_novos, temas_existentes):
    """Filtra temas que já existem ou são muito similares aos existentes."""
    temas_filtrados = []
    
    for tema_obj in temas_novos:
        if not isinstance(tema_obj, dict):
            continue
            
        tema_nome = tema_obj.get('tema', tema_obj.get('Tema', ''))
        if not tema_nome:
            continue
        
        tema_normalizado = tema_nome.lower().strip()
        
        # Verifica se é duplicata exata
        if tema_normalizado in temas_existentes:
            print(f"⚠️ Tema duplicado ignorado: '{tema_nome}'")
            continue
        
        # Verifica se é similar a algum tema existente
        eh_similar = False
        for tema_existente in temas_existentes:
            if temas_sao_similares(tema_nome, tema_existente):
                print(f"⚠️ Tema similar ignorado: '{tema_nome}' (similar a '{tema_existente}')")
                eh_similar = True
                break
        
        if not eh_similar:
            temas_filtrados.append(tema_obj)
    
    return temas_filtrados


def salvar_planilha(response_text):
    """Adiciona os temas extraídos à planilha Excel existente ou cria uma nova, evitando duplicatas."""
    try:
        # Extrai JSON do texto (remove markdown code blocks se existirem)
        texto_limpo = response_text.strip()
        if '```' in texto_limpo:
            inicio = texto_limpo.find('{')
            fim = texto_limpo.rfind('}') + 1
            if inicio != -1 and fim > inicio:
                texto_limpo = texto_limpo[inicio:fim]
        
        dados_json = json.loads(texto_limpo)
        temas = dados_json.get('top_themes', [])
        
        if not temas:
            print("⚠️ Nenhum tema encontrado no JSON.")
            return
        
        arquivo_planilha = 'planilha_temas.xlsx'
        
        # Lê temas existentes para evitar duplicatas
        temas_existentes = ler_temas_existentes(arquivo_planilha)
        print(f"📋 Encontrados {len(temas_existentes)} tema(s) existente(s) na planilha.")
        
        # Filtra temas repetidos
        temas_filtrados = filtrar_temas_repetidos(temas, temas_existentes)
        
        if not temas_filtrados:
            print("⚠️ Todos os temas gerados já existem na planilha. Nenhum novo tema será adicionado.")
            return
        
        print(f"✅ {len(temas_filtrados)} tema(s) novo(s) serão adicionados (de {len(temas)} tema(s) gerado(s)).")
        
        headers = ['Tema', 'Descrição', 'Relevância', 'Roteiro', 'Video Pronto', 'Video Postado', 'Data']
        
        # Verifica se o arquivo existe
        if os.path.exists(arquivo_planilha):
            workbook = load_workbook(arquivo_planilha)
            worksheet = workbook.active
            
            # Garante que os cabeçalhos existam (atualiza se necessário)
            if worksheet.max_row == 0 or worksheet.cell(1, 1).value != 'Tema':
                for col, header in enumerate(headers, start=1):
                    cell = worksheet.cell(1, col)
                    cell.value = header
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
            
            # Encontra a próxima linha vazia
            proxima_linha = worksheet.max_row + 1
        else:
            # Cria nova planilha
            workbook = Workbook()
            worksheet = workbook.active
            
            # Adiciona cabeçalhos
            for col, header in enumerate(headers, start=1):
                cell = worksheet.cell(1, col)
                cell.value = header
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
            
            proxima_linha = 2
        
        # Adiciona apenas os novos temas (já filtrados)
        for tema in temas_filtrados:
            if isinstance(tema, dict):
                tema_nome = tema.get('tema', tema.get('Tema', ''))
                worksheet.cell(proxima_linha, 1, tema_nome)
                worksheet.cell(proxima_linha, 2, tema.get('descricao', tema.get('Descrição', '')))
                worksheet.cell(proxima_linha, 3, tema.get('relevancia', tema.get('Relevância', '')))
                # Deixa Roteiro, Video Pronto, Video Postado e Data em branco
                worksheet.cell(proxima_linha, 4, '')  # Roteiro
                worksheet.cell(proxima_linha, 5, '')  # Video Pronto
                worksheet.cell(proxima_linha, 6, '')  # Video Postado
                worksheet.cell(proxima_linha, 7, '')  # Data
                proxima_linha += 1
        
        # Ajusta largura das colunas
        worksheet.column_dimensions['A'].width = 30
        worksheet.column_dimensions['B'].width = 50
        worksheet.column_dimensions['C'].width = 15
        worksheet.column_dimensions['D'].width = 50  # Roteiro
        worksheet.column_dimensions['E'].width = 15  # Video Pronto
        worksheet.column_dimensions['F'].width = 15  # Video Postado
        worksheet.column_dimensions['G'].width = 12  # Data
        
        workbook.save(arquivo_planilha)
        print(f"\n✅ Planilha atualizada com sucesso! ({len(temas_filtrados)} tema(s) adicionado(s))")
        
    except json.JSONDecodeError as e:
        print(f"❌ Erro ao fazer parse do JSON: {e}")
    except Exception as e:
        print(f"❌ Erro ao salvar planilha: {e}")


def gerar_temas_tiktok_studio(tipo_tema='atualidades', quantidade_temas=3, api_key=None):
    """
    Gera temas usando o TikTok Studio ou a API do Gemini para temas específicos.
    
    Args:
        tipo_tema: 'atualidades', 'terror', 'lenda urbana' ou 'espiritualidade'
        api_key: Chave da API do Gemini. Se None, usa a chave hardcoded.
    
    Returns:
        bool: True se os temas foram gerados com sucesso, False caso contrário.
    """
    # Configura API Key
    if api_key is None:
        api_key = "AIzaSyDZ_6FweRyBza_TuiWQ1W9zgubhfzHqRyY"
    
    if not api_key:
        print("❌ Erro: GEMINI_API_KEY não foi definida.")
        return False
        
    genai.configure(api_key=api_key)
    model = genai.GenerativeModel('gemini-2.5-flash')

    prompt = None
    if tipo_tema == 'lenda urbana':
        print(f"\n✅ Gerando temas de {tipo_tema} via API...")
        prompt = f"""
Você é um criador de conteúdo especializado em lendas urbanas e mistérios.
Sua missão é gerar exatamente {quantidade_temas} TEMAS para vídeos virais de TikTok, 
onde cada vídeo explora uma lenda urbana brasileira de forma sombria e envolvente.

🎯 REGRAS PRINCIPAIS:
- Cada tema representa UM vídeo.
- As lendas devem ser brasileiras (ex: Corpo Seco, Chupa-cabra, Loiras do Banheiro, etc.).
- O tom deve ser de suspense, mistério e um pouco assustador.
- Sempre comece com um **gancho forte** no estilo: “Você já ouviu falar da lenda do...?”
- A **descrição** deve conter de 5 a 8 linhas, detalhando a origem da lenda, os eventos principais e o mistério que a cerca.
- Finalize cada descrição com uma pergunta que incentive o engajamento, tipo: “Você teria coragem de...?”

🪶 EXEMPLO DE SAÍDA IDEAL:
{{
  "top_themes": [
    {{
      "tema": "A Lenda do Corpo Seco",
      "descricao": "Dizem que em Minas Gerais, um homem tão cruel em vida foi rejeitado pela terra e pelo céu. Seu corpo, agora seco e amaldiçoado, vaga pelas estradas assombrando viajantes. A lenda conta que ele ataca quem passa à noite, sugando sua energia vital para tentar reviver. Você teria coragem de passar por uma estrada deserta à noite?",
      "relevancia": "alta"
    }},
    {{
      "tema": "O Mistério da Loira do Banheiro",
      "descricao": "Em escolas de todo o Brasil, uma lenda arrepia os alunos. Uma jovem loira, morta tragicamente no banheiro da escola, assombra o local. Dizem que se você chamar seu nome três vezes no espelho, ela aparece. O que ela quer? Ninguém sabe ao certo, mas seu espírito parece buscar vingança ou apenas companhia. Você se atreveria a invocá-la?",
      "relevancia": "alta"
    }}
  ]
}}

⚠️ FORMATO OBRIGATÓRIO:
Retorne SOMENTE o JSON acima, sem texto extra, explicações ou markdown.
"""
    elif tipo_tema == 'espiritualidade':
        print(f"\n✅ Gerando temas de {tipo_tema} via API...")
        prompt = f"""
Você é um criador de conteúdo especializado em espiritualidade e mistérios do cotidiano.
Sua missão é gerar exatamente {quantidade_temas} TEMAS para vídeos virais de TikTok, 
onde cada vídeo revela os PODERES SOBRENATURAIS, ENERGIAS ESPIRITUAIS ou DONS OCULTOS 
associados a QUATRO sobrenomes comuns no Brasil.

🎯 REGRAS PRINCIPAIS:
- Cada tema representa UM vídeo.
- Cada vídeo deve conter exatamente **4 sobrenomes diferentes**.
- Todos os sobrenomes devem ser **populares no Brasil** — exemplos: Silva, Souza, Alves, Costa, Oliveira, Rocha, Nascimento, Lima, Carvalho, Gomes, Melo, Martins, Falcão, Portela, Amaral, etc.
- Misture significados místicos, espirituais e simbólicos (ancestralidade, intuição, proteção, dons ocultos, maldições antigas, etc.).
- O tom deve ser **misterioso, documental e espiritual**, como se fosse uma revelação antiga.
- Sempre comece com um **gancho chamativo** no estilo: “Você sabia que alguns sobrenomes escondem poderes espirituais há gerações?”
- A **descrição** deve conter de 5 a 8 linhas, descrevendo os 4 sobrenomes e seus dons/poderes.
- Finalize cada descrição com uma chamada leve, tipo: “Manda esse vídeo pra alguém com um desses nomes.”

🪶 EXEMPLO DE SAÍDA IDEAL:
{{
  "top_themes": [
    {{
      "tema": "Os Sobrenomes Que Herdaram Dons Ocultos",
      "descricao": "Você sabia que alguns sobrenomes carregam energia espiritual há séculos? Os Silva são guardiões naturais — sentem presenças e têm o dom da proteção. Os Souza possuem intuição poderosa e corpo fechado contra o mal. Já os Amaral vêm de antigas linhagens judaicas ligadas à sabedoria mística. E os Oliveira, conectados à árvore sagrada, trazem paz e equilíbrio por onde passam. Manda pra alguém com um desses nomes.",
      "relevancia": "alta"
    }},
    {{
      "tema": "Sobrenomes Que Nascem Com Poder",
      "descricao": "Dizem que os Lima têm o dom da cura espiritual, capazes de transformar ambientes com a energia das mãos. Os Rocha carregam firmeza e atraem força ancestral. Os Nascimento são almas de recomeço — renascem das cinzas sempre mais fortes. E os Gomes têm magnetismo natural, atraindo o que desejam com o poder do pensamento. Manda esse vídeo pra quem tem um desses nomes.",
      "relevancia": "alta"
    }}
  ]
}}

⚠️ FORMATO OBRIGATÓRIO:
Retorne SOMENTE o JSON acima, sem texto extra, explicações ou markdown.

💡 TOM E ESTILO:
- Mistério + Espiritualidade + Curiosidade
- Linguagem emocional, mas leve
- Sem religiosidade direta
- Estilo ideal para vídeos narrados no TikTok
"""


    if prompt:
        try:
            print("\n🤖 Enviando prompt para o Gemini...")
            response = model.generate_content(prompt)
            
            print("\n--- Resposta do Gemini ---")
            print(response.text)
            print("--- Fim da Resposta ---\n")

            salvar_planilha(response.text)
            return True

        except Exception as e:
            if "API key" in str(e):
                print("❌ Erro de autenticação com a API do Gemini. Verifique sua API Key.")
            else:
                print(f"❌ Erro ao usar a API do Gemini: {e}")
            return False

    # Lógica existente para 'atualidades' e 'terror'
    if tipo_tema == 'atualidades':
        numero_tabs = 16
    elif tipo_tema == 'terror':
        numero_tabs = 28
    else:
        print(f"⚠️ Tipo de tema desconhecido: {tipo_tema}. Usando 'atualidades'.")
        tipo_tema = 'atualidades'
        numero_tabs = 16
    
    print(f"\n✅ Buscando temas de {tipo_tema} no TikTok Studio...")
    
    try:
        # Abre TikTok Studio
        url = 'https://www.tiktok.com/tiktokstudio/inspiration'
        webbrowser.open(url)
        time.sleep(10)
        pyautogui.click(x=1212, y=229)
        pyautogui.sleep(4)
        pyautogui.click(x=1184, y=522)
        pyautogui.sleep(5)

        # Navega até o conteúdo
        for _ in range(numero_tabs):
            pyautogui.press('tab')

        pyautogui.press('enter')
        time.sleep(2)

        # Copia o conteúdo
        pyautogui.hotkey('ctrl', 'a')
        pyautogui.hotkey('ctrl', 'c')
        time.sleep(0.5)

        conteudo = pyperclip.paste()
        
        prompt_tiktok = f"""Analise o texto a seguir (copiado da página de Inspiração do TikTok Studio) e identifique exatamente os {quantidade_temas} TÓPICOS mais relevantes.
(Resto do prompt do TikTok Studio...)
"""

        print("\n🤖 Enviando texto para análise do Gemini...")
        response = model.generate_content(prompt_tiktok)
        
        print("\n--- Análise do Gemini ---")
        print(response.text)
        print("--- Fim da Análise ---\n")

        salvar_planilha(response.text)
        return True

    except Exception as e:
        if "API key" in str(e):
            print("❌ Erro de autenticação com a API do Gemini. Verifique sua API Key.")
        else:
            print(f"❌ Erro ao usar a API do Gemini: {e}")
        return False


# Código principal para execução direta do script
if __name__ == "__main__":
    # Pergunta qual tipo de tema o usuário deseja
    print("\n🔍 Escolha o tipo de tema:")
    print("1 - Atualidades (via TikTok Studio)")
    print("2 - Terror (via TikTok Studio)")
    print("3 - Lenda Urbana (via API Gemini)")
    print("4 - Espiritualidade (via API Gemini)")
    escolha = input("Digite o número da opção: ").strip()

    while escolha not in ['1', '2', '3', '4']:
        print("⚠️ Opção inválida. Digite 1, 2, 3 ou 4.")
        escolha = input("Digite o número da opção: ").strip()

    if escolha == '1':
        tipo_tema = "atualidades"
    elif escolha == '2':
        tipo_tema = "terror"
    elif escolha == '3':
        tipo_tema = "lenda urbana"
    else:
        tipo_tema = "espiritualidade"

    quantidade_temas_str = input(f"Digite a quantidade de temas de '{tipo_tema}' a serem gerados (padrão: 3): ").strip()
    if not quantidade_temas_str:
        quantidade_temas_str = "3" # Valor padrão

    while not quantidade_temas_str.isdigit() or int(quantidade_temas_str) <= 0:
        print("⚠️ Quantidade inválida. Digite um número inteiro positivo.")
        quantidade_temas_str = input(f"Digite a quantidade de temas de '{tipo_tema}' a serem gerados (padrão: 3): ").strip()
        if not quantidade_temas_str:
            quantidade_temas_str = "3" # Valor padrão

    gerar_temas_tiktok_studio(tipo_tema, quantidade_temas=int(quantidade_temas_str))
    print("\n✅ Processo concluído.")

