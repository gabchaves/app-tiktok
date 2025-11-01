import pyautogui as teclado
import pyperclip
from openpyxl import load_workbook
from datetime import datetime
import os

# Busca o último roteiro que tem roteiro preenchido mas Video Pronto ainda vazio
def buscar_ultimo_roteiro():
    planilha_path = 'planilha_temas.xlsx'
    
    if not os.path.exists(planilha_path):
        print("❌ Planilha não encontrada!")
        return None, None
    
    try:
        workbook = load_workbook(planilha_path)
        worksheet = workbook.active
        
        # Procura de baixo para cima o último roteiro que tem roteiro mas não tem Video Pronto OK
        ultima_linha = None
        ultimo_roteiro = None
        
        for row in range(worksheet.max_row, 1, -1):  # Vai da última linha até a linha 2
            roteiro = worksheet.cell(row, 4).value  # Coluna D = Roteiro
            video_pronto = worksheet.cell(row, 5).value  # Coluna E = Video Pronto
            
            # Se tem roteiro e ainda não está marcado como Video Pronto OK
            if roteiro and (not video_pronto or str(video_pronto).strip().upper() != 'OK'):
                ultima_linha = row
                ultimo_roteiro = roteiro
                break
        
        if ultima_linha and ultimo_roteiro:
            print(f"✅ Roteiro encontrado na linha {ultima_linha}!")
            return ultimo_roteiro, ultima_linha, workbook, worksheet
        else:
            print("⚠️ Nenhum roteiro pendente encontrado!")
            return None, None, None, None
            
    except Exception as e:
        print(f"❌ Erro ao ler planilha: {e}")
        return None, None, None, None

# Marca Video Pronto como OK e preenche a data
def marcar_video_pronto(workbook, worksheet, linha):
    try:
        # Coluna E = Video Pronto, Coluna G = Data
        worksheet.cell(row=linha, column=5, value="OK")
        worksheet.cell(row=linha, column=7, value=datetime.now().strftime("%Y-%m-%d"))
        workbook.save('planilha_temas.xlsx')
        print(f"✅ Video Pronto marcado como OK na linha {linha} com data {datetime.now().strftime('%Y-%m-%d')}!")
    except Exception as e:
        print(f"❌ Erro ao atualizar planilha: {e}")

# Busca o roteiro
roteiro, linha, workbook, worksheet = buscar_ultimo_roteiro()

if roteiro:
    # Copia o roteiro para a área de transferência
    pyperclip.copy(roteiro)
    print("📋 Roteiro copiado para área de transferência!")
    
    # Abre o aplicativo CapCut
    teclado.hotkey('win', 'r')
    teclado.typewrite('https://www.capcut.com/ai-creator/start')
    teclado.press('enter')
    teclado.sleep(5)
    teclado.click(x=298, y=487)
    teclado.sleep(5)
    teclado.click(x=1043, y=492)
    teclado.sleep(5)
    teclado.click(x=1080, y=553)
    teclado.sleep(5)
    teclado.click(x=1022, y=610)
    teclado.sleep(5)
    
    # Cola o roteiro (Ctrl+V)
    teclado.hotkey('ctrl', 'v')
    teclado.sleep(2)  # Aguarda um pouco após colar
    
    # Marca Video Pronto como OK e preenche a data
    marcar_video_pronto(workbook, worksheet, linha)
    
    teclado.press('tab')
    teclado.press('tab')
    teclado.press('enter')
    teclado.sleep(120)
    teclado.click(x=1042, y=153)
    

    print("✅ Processo concluído!")
else:
    print("❌ Não foi possível continuar. Nenhum roteiro encontrado.")

