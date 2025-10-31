import openpyxl
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill
import os
import google.generativeai as genai


def abrir_planilha():
    """Abre a planilha de temas e garante que todas as colunas necessárias existam."""
    arquivo_planilha = 'planilha_temas.xlsx'
    
    try:
        if os.path.exists(arquivo_planilha):
            workbook = load_workbook(arquivo_planilha)
            worksheet = workbook.active
            
            # Garante que todas as colunas existam
            headers = ['Tema', 'Descrição', 'Relevância', 'Roteiro', 'Video Pronto', 'Video Postado', 'Data']
            
            # Verifica e atualiza cabeçalhos se necessário
            if worksheet.max_row == 0 or worksheet.cell(1, 1).value != 'Tema':
                for col, header in enumerate(headers, start=1):
                    cell = worksheet.cell(1, col)
                    cell.value = header
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
            
            # Preenche colunas vazias nas linhas existentes (exceto cabeçalho)
            for row in range(2, worksheet.max_row + 1):
                # Garante que as colunas extras estejam vazias se não tiverem valor
                if worksheet.cell(row, 4).value is None:
                    worksheet.cell(row, 4, '')  # Roteiro
                if worksheet.cell(row, 5).value is None:
                    worksheet.cell(row, 5, '')  # Video Pronto
                if worksheet.cell(row, 6).value is None:
                    worksheet.cell(row, 6, '')  # Video Postado
                if worksheet.cell(row, 7).value is None:
                    worksheet.cell(row, 7, '')  # Data
            
            # Ajusta largura das colunas
            worksheet.column_dimensions['A'].width = 30
            worksheet.column_dimensions['B'].width = 50
            worksheet.column_dimensions['C'].width = 15
            worksheet.column_dimensions['D'].width = 50  # Roteiro
            worksheet.column_dimensions['E'].width = 15  # Video Pronto
            worksheet.column_dimensions['F'].width = 15  # Video Postado
            worksheet.column_dimensions['G'].width = 12  # Data
            
            workbook.save(arquivo_planilha)
            print(f"✅ Planilha '{arquivo_planilha}' aberta e verificada com sucesso!")
            print(f"📊 Total de linhas: {worksheet.max_row - 1} tema(s)")
            
            return workbook, worksheet
        else:
            print(f"⚠️ Planilha '{arquivo_planilha}' não encontrada.")
            print("💡 Execute o App.py primeiro para criar a planilha.")
            return None, None
            
    except Exception as e:
        print(f"❌ Erro ao abrir planilha: {e}")
        return None, None


def gerar_roteiro(tema, descricao):
    """Gera um roteiro para vídeo usando Gemini AI."""
    GEMINI_API_KEY = "AIzaSyDZ_6FweRyBza_TuiWQ1W9zgubhfzHqRyY"
    
    if not GEMINI_API_KEY:
        print("❌ Erro: GEMINI_API_KEY não foi definida.")
        return None
    
    try:
        genai.configure(api_key=GEMINI_API_KEY)
        model = genai.GenerativeModel('gemini-2.5-flash')
        
        prompt = f"""Crie um roteiro completo e detalhado para um vídeo do TikTok sobre o seguinte tema:

Tema: {tema}
Descrição: {descricao}

O roteiro deve incluir:
- Introdução chamativa (primeiros 3 segundos)
- Desenvolvimento do conteúdo (meio do vídeo)
- Chamada para ação ou conclusão (final)

Seja criativo, engajador e adequado para o formato TikTok (vídeos curtos e dinâmicos).
Retorne apenas o roteiro, sem formatação markdown ou códigos."""
        
        print(f"\n🤖 Gerando roteiro para: {tema}...")
        response = model.generate_content(prompt)
        roteiro = response.text.strip()
        
        print("✅ Roteiro gerado com sucesso!")
        return roteiro
        
    except Exception as e:
        if "API key" in str(e):
            print("❌ Erro de autenticação com a API do Gemini. Verifique sua API Key.")
        else:
            print(f"❌ Erro ao gerar roteiro: {e}")
        return None


def processar_primeiro_tema(workbook, worksheet):
    """Pega o primeiro tema da planilha, gera roteiro e salva."""
    # Verifica se há temas na planilha (linha 2 em diante)
    if worksheet.max_row < 2:
        print("⚠️ Nenhum tema encontrado na planilha.")
        return None
    
    # Pega o primeiro tema (linha 2)
    tema = worksheet.cell(2, 1).value  # Coluna A - Tema
    descricao = worksheet.cell(2, 2).value  # Coluna B - Descrição
    roteiro_atual = worksheet.cell(2, 4).value  # Coluna D - Roteiro
    
    if not tema:
        print("⚠️ O primeiro tema está vazio.")
        return None
    
    # Verifica se já tem roteiro
    if roteiro_atual and roteiro_atual.strip():
        print(f"⚠️ O tema '{tema}' já possui um roteiro.")
        resposta = input("Deseja gerar um novo roteiro? (s/n): ").lower()
        if resposta != 's':
            print("Operação cancelada.")
            return roteiro_atual
    
    print(f"\n📋 Processando primeiro tema:")
    print(f"   Tema: {tema}")
    print(f"   Descrição: {descricao or 'N/A'}")
    
    # Gera o roteiro
    roteiro = gerar_roteiro(tema, descricao or "")
    
    if roteiro:
        # Salva na planilha (coluna D - linha 2)
        worksheet.cell(2, 4, roteiro)
        
        # Salva o arquivo
        arquivo_planilha = 'planilha_temas.xlsx'
        workbook.save(arquivo_planilha)
        
        print(f"\n✅ Roteiro salvo na planilha!")
        print(f"\n--- Roteiro Gerado ---")
        print(roteiro)
        print("--- Fim do Roteiro ---\n")
        
        return roteiro
    else:
        print("❌ Não foi possível gerar o roteiro.")
        return None


if __name__ == "__main__":
    print("🎬 Video Creator - Gerando roteiro para primeiro tema...\n")
    workbook, worksheet = abrir_planilha()
    
    if workbook and worksheet:
        roteiro = processar_primeiro_tema(workbook, worksheet)
        if roteiro:
            print("\n✅ Processo concluído com sucesso!")
        else:
            print("\n⚠️ Não foi possível processar o primeiro tema.")
    else:
        print("\n❌ Não foi possível abrir a planilha.")

