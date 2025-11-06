import os
from openpyxl import load_workbook, Workbook
from Roteiro_generator import RoteiroGenerator
from Tema_generator import gerar_temas_tiktok_studio
import pyautogui as teclado
import pyperclip
from datetime import datetime
import google.generativeai as genai
import json
from openpyxl.styles import Font, PatternFill
import time

# API Key do Gemini
GEMINI_API_KEY = "AIzaSyDZ_6FweRyBza_TuiWQ1W9zgubhfzHqRyY"
PLANILHA_PATH = 'planilha_temas.xlsx'


def contar_temas_disponiveis():
    """Conta quantos temas existem na planilha (com Tema preenchido)."""
    if not os.path.exists(PLANILHA_PATH):
        return 0
    
    try:
        workbook = load_workbook(PLANILHA_PATH)
        worksheet = workbook.active
        
        count = 0
        for row in range(2, worksheet.max_row + 1):
            tema = worksheet.cell(row, 1).value
            if tema and str(tema).strip():
                count += 1
        
        return count
    except Exception as e:
        print(f"❌ Erro ao contar temas: {e}")
        return 0


def contar_roteiros_disponiveis():
    """Conta quantos roteiros existem na planilha (com Roteiro preenchido e Video Pronto não OK)."""
    if not os.path.exists(PLANILHA_PATH):
        return 0
    
    try:
        workbook = load_workbook(PLANILHA_PATH)
        worksheet = workbook.active
        
        count = 0
        for row in range(2, worksheet.max_row + 1):
            roteiro = worksheet.cell(row, 4).value
            video_pronto = worksheet.cell(row, 5).value
            
            if roteiro and str(roteiro).strip():
                # Não conta se já está marcado como OK
                if not video_pronto or str(video_pronto).strip().upper() != 'OK':
                    count += 1
        
        return count
    except Exception as e:
        print(f"❌ Erro ao contar roteiros: {e}")
        return 0


def contar_videos_prontos():
    """Conta quantos vídeos já estão prontos (com Video Pronto = OK)."""
    if not os.path.exists(PLANILHA_PATH):
        return 0
    
    try:
        workbook = load_workbook(PLANILHA_PATH)
        worksheet = workbook.active
        
        count = 0
        for row in range(2, worksheet.max_row + 1):
            video_pronto = worksheet.cell(row, 5).value
            if video_pronto and str(video_pronto).strip().upper() == 'OK':
                count += 1
        
        return count
    except Exception as e:
        print(f"❌ Erro ao contar vídeos prontos: {e}")
        return 0


def contar_temas_sem_roteiro():
    """Conta quantos temas ainda não têm roteiro gerado."""
    if not os.path.exists(PLANILHA_PATH):
        return 0
    
    try:
        workbook = load_workbook(PLANILHA_PATH)
        worksheet = workbook.active
        
        count = 0
        for row in range(2, worksheet.max_row + 1):
            tema = worksheet.cell(row, 1).value
            roteiro = worksheet.cell(row, 4).value
            
            # Conta apenas temas que existem mas não têm roteiro
            if tema and str(tema).strip():
                if not roteiro or not str(roteiro).strip():
                    count += 1
        
        return count
    except Exception as e:
        print(f"❌ Erro ao contar temas sem roteiro: {e}")
        return 0


def ler_temas_existentes():
    """Lê todos os temas já existentes na planilha para evitar duplicatas."""
    temas_existentes = set()
    
    if not os.path.exists(PLANILHA_PATH):
        return temas_existentes
    
    try:
        workbook = load_workbook(PLANILHA_PATH)
        worksheet = workbook.active
        
        for row in range(2, worksheet.max_row + 1):
            tema = worksheet.cell(row, 1).value
            if tema and isinstance(tema, str):
                tema_normalizado = tema.lower().strip()
                temas_existentes.add(tema_normalizado)
        
        return temas_existentes
    except Exception as e:
        print(f"⚠️ Erro ao ler temas existentes: {e}")
        return temas_existentes


def temas_sao_similares(tema1, tema2):
    """Verifica se dois temas são muito similares."""
    t1 = tema1.lower().strip()
    t2 = tema2.lower().strip()
    
    if t1 == t2:
        return True
    
    palavras_t1 = set(t1.split())
    palavras_t2 = set(t2.split())
    
    palavras_principais_t1 = {p for p in palavras_t1 if len(p) > 3}
    palavras_principais_t2 = {p for p in palavras_t2 if len(p) > 3}
    
    if palavras_principais_t1 and palavras_principais_t2:
        palavras_comuns = palavras_principais_t1 & palavras_principais_t2
        todas_palavras = palavras_principais_t1 | palavras_principais_t2
        if todas_palavras and len(palavras_comuns) / len(todas_palavras) > 0.7:
            return True
    
    return False


def filtrar_temas_repetidos(temas_novos, temas_existentes):
    """Filtra temas que já existem ou são muito similares."""
    temas_filtrados = []
    
    for tema_obj in temas_novos:
        if not isinstance(tema_obj, dict):
            continue
            
        tema_nome = tema_obj.get('tema', tema_obj.get('Tema', ''))
        if not tema_nome:
            continue
        
        tema_normalizado = tema_nome.lower().strip()
        
        if tema_normalizado in temas_existentes:
            continue
        
        eh_similar = False
        for tema_existente in temas_existentes:
            if temas_sao_similares(tema_nome, tema_existente):
                eh_similar = True
                break
        
        if not eh_similar:
            temas_filtrados.append(tema_obj)
    
    return temas_filtrados


def gerar_temas_automaticos(quantidade):
    """Gera temas automaticamente usando a API do Gemini."""
    try:
        genai.configure(api_key=GEMINI_API_KEY)
        model = genai.GenerativeModel('gemini-2.5-flash')
        
        # Calcula quantos lotes de 3 temas precisamos (já que a API gera 3 por vez)
        lotes = (quantidade + 2) // 3
        
        temas_existentes = ler_temas_existentes()
        temas_total = []
        
        for lote in range(lotes):
            print(f"\n🤖 Gerando lote {lote + 1} de {lotes}...")
            
            prompt = """Gere 3 temas relevantes para criação de vídeos virais no TikTok, Instagram Reels e YouTube Shorts.
Os temas devem ser sobre mistérios, histórias reais, teorias interessantes, fatos chocantes, ou eventos misteriosos.
Retorne a resposta em formato JSON, com a seguinte estrutura:
{
  "top_themes": [
    {"tema": "nome do tema", "descricao": "explicação detalhada", "relevancia": "alta|média|baixa"},
    {"tema": "nome do tema", "descricao": "explicação detalhada", "relevancia": "alta|média|baixa"},
    {"tema": "nome do tema", "descricao": "explicação detalhada", "relevancia": "alta|média|baixa"}
  ]
}

📋 INSTRUÇÕES PARA AS DESCRIÇÕES:
Cada descrição deve ser um texto de 2 a 4 frases que contenha:
- Um gancho inicial forte (elemento chocante, curioso ou emocional)
- Detalhes específicos que ajudem na criação do roteiro (fatos, números, eventos, personagens, locais, datas)
- Indicação de elementos visuais ou emocionais que tornam o tema viral (mistério, tensão, reviravolta, emoção, curiosidade)
- Contexto suficiente para um roteirista criar um vídeo envolvente

💡 Exemplo de descrição BOA:
"Em 2015, um avião desapareceu sem deixar rastros no meio do oceano. A investigação revelou que todos os passageiros sumiram antes do pouso, deixando apenas objetos pessoais. O mistério nunca foi resolvido, alimentando teorias sobre dimensões paralelas e sequestros extraterrestres. Este caso desperta medo, curiosidade e debate, elementos perfeitos para um vídeo viral."

❌ Exemplo de descrição RUIM (muito genérica):
"Um avião desapareceu e gerou mistério sobre o que aconteceu."

Retorne APENAS o JSON, sem texto adicional."""
            
            response = model.generate_content(prompt)
            
            # Extrai JSON do texto
            texto_limpo = response.text.strip()
            if '```' in texto_limpo:
                inicio = texto_limpo.find('{')
                fim = texto_limpo.rfind('}') + 1
                if inicio != -1 and fim > inicio:
                    texto_limpo = texto_limpo[inicio:fim]
            
            dados_json = json.loads(texto_limpo)
            temas_lote = dados_json.get('top_themes', [])
            temas_total.extend(temas_lote)
            
            # Atualiza temas existentes para evitar duplicatas
            for tema in temas_lote:
                if isinstance(tema, dict):
                    tema_nome = tema.get('tema', tema.get('Tema', ''))
                    if tema_nome:
                        temas_existentes.add(tema_nome.lower().strip())
            
            time.sleep(1)  # Pausa entre requisições
        
        # Filtra duplicatas
        temas_filtrados = filtrar_temas_repetidos(temas_total, ler_temas_existentes())
        
        return temas_filtrados
    
    except Exception as e:
        print(f"❌ Erro ao gerar temas: {e}")
        return []


def salvar_temas_na_planilha(temas):
    """Salva temas na planilha."""
    if not temas:
        return
    
    try:
        headers = ['Tema', 'Descrição', 'Relevância', 'Roteiro', 'Video Pronto', 'Video Postado', 'Data']
        
        if os.path.exists(PLANILHA_PATH):
            workbook = load_workbook(PLANILHA_PATH)
            worksheet = workbook.active
            
            if worksheet.max_row == 0 or worksheet.cell(1, 1).value != 'Tema':
                for col, header in enumerate(headers, start=1):
                    cell = worksheet.cell(1, col)
                    cell.value = header
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
            
            proxima_linha = worksheet.max_row + 1
        else:
            workbook = Workbook()
            worksheet = workbook.active
            
            for col, header in enumerate(headers, start=1):
                cell = worksheet.cell(1, col)
                cell.value = header
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
            
            proxima_linha = 2
        
        for tema in temas:
            if isinstance(tema, dict):
                tema_nome = tema.get('tema', tema.get('Tema', ''))
                worksheet.cell(proxima_linha, 1, tema_nome)
                worksheet.cell(proxima_linha, 2, tema.get('descricao', tema.get('Descrição', '')))
                worksheet.cell(proxima_linha, 3, tema.get('relevancia', tema.get('Relevância', '')))
                worksheet.cell(proxima_linha, 4, '')  # Roteiro
                worksheet.cell(proxima_linha, 5, '')  # Video Pronto
                worksheet.cell(proxima_linha, 6, '')  # Video Postado
                worksheet.cell(proxima_linha, 7, '')  # Data
                proxima_linha += 1
        
        worksheet.column_dimensions['A'].width = 30
        worksheet.column_dimensions['B'].width = 50
        worksheet.column_dimensions['C'].width = 15
        worksheet.column_dimensions['D'].width = 50
        worksheet.column_dimensions['E'].width = 15
        worksheet.column_dimensions['F'].width = 15
        worksheet.column_dimensions['G'].width = 12
        
        workbook.save(PLANILHA_PATH)
        print(f"✅ {len(temas)} tema(s) adicionado(s) à planilha!")
    
    except Exception as e:
        print(f"❌ Erro ao salvar temas: {e}")


def garantir_temas_suficientes(quantidade_necessaria):
    """Garante que há temas suficientes na planilha."""
    temas_sem_roteiro = contar_temas_sem_roteiro()
    
    print(f"\n📊 Status inicial:")
    print(f"   Temas disponíveis (sem roteiro): {temas_sem_roteiro}")
    print(f"   Temas necessários: {quantidade_necessaria}")
    
    if temas_sem_roteiro >= quantidade_necessaria:
        print(f"✅ Temas suficientes!")
        return True
    
    print(f"\n⚠️ Faltam temas. Usando Tema Generator do TikTok Studio...")
    print("\n🔍 Escolha o tipo de tema:")
    print("1 - Atualidades")
    print("2 - Terror")
    escolha = input("Digite o número da opção: ").strip()
    
    while escolha not in ['1', '2']:
        print("⚠️ Opção inválida. Digite 1 para Atualidades ou 2 para Terror.")
        escolha = input("Digite o número da opção: ").strip()
    
    if escolha == '1':
        tipo_tema = "atualidades"
    else:
        tipo_tema = "terror"
    
    # Executa o Tema Generator
    sucesso = gerar_temas_tiktok_studio(tipo_tema, GEMINI_API_KEY)
    
    if sucesso:
        temas_sem_roteiro = contar_temas_sem_roteiro()
        print(f"✅ Agora temos {temas_sem_roteiro} tema(s) disponíveis.")
        
        # Se ainda não tem temas suficientes, tenta gerar mais com a API
        if temas_sem_roteiro < quantidade_necessaria:
            falta = quantidade_necessaria - temas_sem_roteiro
            print(f"\n⚠️ Ainda faltam {falta} tema(s). Gerando temas via API...")
            
            temas_gerados = gerar_temas_automaticos(falta)
            if temas_gerados:
                salvar_temas_na_planilha(temas_gerados)
                temas_sem_roteiro = contar_temas_sem_roteiro()
                print(f"✅ Agora temos {temas_sem_roteiro} tema(s) disponíveis.")
        
        return temas_sem_roteiro >= quantidade_necessaria
    
    return False


def garantir_roteiros_suficientes(quantidade_necessaria):
    """Garante que há roteiros suficientes."""
    roteiros_atual = contar_roteiros_disponiveis()
    
    print(f"\n📊 Status de roteiros:")
    print(f"   Roteiros disponíveis: {roteiros_atual}")
    print(f"   Roteiros necessários: {quantidade_necessaria}")
    
    if roteiros_atual >= quantidade_necessaria:
        print(f"✅ Roteiros suficientes!")
        return True
    
    falta = quantidade_necessaria - roteiros_atual
    print(f"\n⚠️ Faltam {falta} roteiro(s). Gerando roteiros...")
    
    generator = RoteiroGenerator(PLANILHA_PATH)
    gerados = 0
    
    while gerados < falta:
        resultado = generator.processar_primeiro_tema()
        if resultado:
            gerados += 1
            print(f"✅ Roteiro {gerados}/{falta} gerado!")
        else:
            print("❌ Não foi possível gerar mais roteiros.")
            break
    
    roteiros_atual = contar_roteiros_disponiveis()
    print(f"✅ Agora temos {roteiros_atual} roteiro(s) disponíveis.")
    return roteiros_atual >= quantidade_necessaria


def buscar_proximo_roteiro():
    """Busca o próximo roteiro que precisa de vídeo."""
    if not os.path.exists(PLANILHA_PATH):
        return None, None, None, None
    
    try:
        workbook = load_workbook(PLANILHA_PATH)
        worksheet = workbook.active
        
        for row in range(2, worksheet.max_row + 1):
            roteiro = worksheet.cell(row, 4).value
            video_pronto = worksheet.cell(row, 5).value
            
            if roteiro and str(roteiro).strip():
                if not video_pronto or str(video_pronto).strip().upper() != 'OK':
                    return roteiro, row, workbook, worksheet
        
        return None, None, None, None
    except Exception as e:
        print(f"❌ Erro ao buscar roteiro: {e}")
        return None, None, None, None


def marcar_video_pronto(workbook, worksheet, linha):
    """Marca o vídeo como pronto na planilha."""
    try:
        worksheet.cell(row=linha, column=5, value="OK")
        worksheet.cell(row=linha, column=7, value=datetime.now().strftime("%Y-%m-%d"))
        workbook.save(PLANILHA_PATH)
        print(f"✅ Video Pronto marcado como OK na linha {linha}!")
    except Exception as e:
        print(f"❌ Erro ao atualizar planilha: {e}")


def gerar_video():
    """Gera um vídeo usando o Video_generator."""
    roteiro, linha, workbook, worksheet = buscar_proximo_roteiro()
    
    if not roteiro:
        return False
    
    try:
        print(f"\n🎬 Gerando vídeo para roteiro da linha {linha}...")
        
        # Copia o roteiro
        pyperclip.copy(roteiro)
        print("📋 Roteiro copiado para área de transferência!")
        
        # Abre o CapCut
        teclado.hotkey('win', 'r')
        teclado.typewrite('https://www.capcut.com/ai-creator/start')
        teclado.press('enter')
        time.sleep(10)
        teclado.click(x=1092, y=329)
        teclado.hotkey('ctrl', 'a')
        
        # Cola o roteiro
        teclado.hotkey('ctrl', 'v')
        time.sleep(2)
        teclado.click(x=1026, y=383)
        teclado.press('tab')
        teclado.press('enter')
        time.sleep(100)
        teclado.click(x=1042, y=153)
        time.sleep(10)
        
        # Marca como pronto
        marcar_video_pronto(workbook, worksheet, linha)
        print("✅ Vídeo gerado com sucesso!")
        return True
    
    except Exception as e:
        print(f"❌ Erro ao gerar vídeo: {e}")
        return False


def gerar_videos(quantidade):
    """Gera a quantidade especificada de vídeos."""
    print(f"\n🎬 Iniciando geração de {quantidade} vídeo(s)...")
    
    gerados = 0
    while gerados < quantidade:
        print(f"\n--- Gerando vídeo {gerados + 1}/{quantidade} ---")
        
        if gerar_video():
            gerados += 1
        else:
            print("⚠️ Não foi possível gerar mais vídeos.")
            break
    
    print(f"\n✅ Processo concluído! {gerados} vídeo(s) gerado(s).")


def main():
    print("=" * 60)
    print("🎬 APP DE GERAÇÃO DE VÍDEOS - TIKTOK")
    print("=" * 60)
    
    try:
        quantidade = int(input("\n📹 Quantos vídeos você deseja gerar? "))
        
        if quantidade <= 0:
            print("❌ Por favor, informe um número maior que zero.")
            return
        
        print(f"\n🚀 Iniciando processo para gerar {quantidade} vídeo(s)...")
        
        # 1. Garantir temas suficientes
        if not garantir_temas_suficientes(quantidade):
            print("❌ Não foi possível garantir temas suficientes.")
            return
        
        # 2. Garantir roteiros suficientes
        if not garantir_roteiros_suficientes(quantidade):
            print("❌ Não foi possível garantir roteiros suficientes.")
            return
        
        # 3. Gerar vídeos
        gerar_videos(quantidade)
        
        # Status final
        temas_sem_roteiro = contar_temas_sem_roteiro()
        roteiros_total = contar_roteiros_disponiveis()
        videos_prontos = contar_videos_prontos()
        
        print("\n" + "=" * 60)
        print("📊 STATUS FINAL:")
        print(f"   Temas sem roteiro: {temas_sem_roteiro}")
        print(f"   Roteiros disponíveis: {roteiros_total}")
        print(f"   Vídeos prontos: {videos_prontos}")
        print("=" * 60)
        
    except ValueError:
        print("❌ Por favor, informe um número válido.")
    except KeyboardInterrupt:
        print("\n\n⚠️ Processo interrompido pelo usuário.")
    except Exception as e:
        print(f"\n❌ Erro inesperado: {e}")


if __name__ == "__main__":
    main()

