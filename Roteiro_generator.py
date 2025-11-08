from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
import os
import google.generativeai as genai
from datetime import datetime

# API Key do Gemini
GEMINI_API_KEY = "AIzaSyDZ_6FweRyBza_TuiWQ1W9zgubhfzHqRyY"

class RoteiroGenerator:
    def __init__(self, planilha_path='planilha_temas.xlsx'):
        self.planilha_path = planilha_path
        self.workbook = None
        self.worksheet = None

    def _abrir_planilha(self):
        """Abre a planilha de temas e garante que todas as colunas necessárias existam."""
        try:
            if os.path.exists(self.planilha_path):
                self.workbook = load_workbook(self.planilha_path)
                self.worksheet = self.workbook.active
                
                headers = ['Tema', 'Descrição', 'Relevância', 'Roteiro', 'Video Pronto', 'Video Postado', 'Data']
                
                if self.worksheet.max_row == 0 or self.worksheet.cell(1, 1).value != 'Tema':
                    for col, header in enumerate(headers, start=1):
                        cell = self.worksheet.cell(1, col)
                        cell.value = header
                        cell.font = Font(bold=True)
                        cell.fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
                
                self.worksheet.column_dimensions['A'].width = 30
                self.worksheet.column_dimensions['B'].width = 50
                self.worksheet.column_dimensions['C'].width = 15
                self.worksheet.column_dimensions['D'].width = 50
                self.worksheet.column_dimensions['E'].width = 15
                self.worksheet.column_dimensions['F'].width = 15
                self.worksheet.column_dimensions['G'].width = 12
                
                self.workbook.save(self.planilha_path)
                print(f"✅ Planilha '{self.planilha_path}' aberta e verificada com sucesso!")
                print(f"📊 Total de linhas: {self.worksheet.max_row - 1} tema(s)")
                return True
            else:
                print(f"⚠️ Planilha '{self.planilha_path}' não encontrada.")
                print("💡 Execute o Tema_generator.py primeiro para criar a planilha.")
                return False
                
        except Exception as e:
            print(f"❌ Erro ao abrir planilha: {e}")
            return False

    def _gerar_roteiro(self, tema, descricao):
        try:
            genai.configure(api_key=GEMINI_API_KEY)
            model = genai.GenerativeModel('gemini-2.5-flash')
            
            prompt = f"""Você é um roteirista especializado em vídeos virais curtos para TikTok, Instagram Reels e YouTube Shorts. Sua missão é criar roteiros com alto potencial de retenção e engajamento, baseados no estilo narrativo de vídeos misteriosos, jornalísticos ou emocionais, com estrutura de impacto, ritmo fluido e linguagem natural.

Crie um roteiro completo e detalhado para um vídeo do TikTok sobre o seguinte tema:

Tema: {tema}
Descrição: {descricao}

💡 Diretrizes obrigatórias:
1. **Gancho inicial (0 a 3 segundos):**
   - Comece com uma frase CHOCANTE, curiosa ou provocativa.
   - Pode usar estruturas como:
     - “Você sabia que…?”
     - “Pouca gente sabe disso, mas…”
     - “O que eu vou te contar agora é real… e assustador.”
     - “Essa história parece ficção, mas aconteceu de verdade.”
   - Deve prender o público já na primeira frase.

2. **Desenvolvimento (meio do vídeo):**
   - Conte a história de forma fluida, como se fosse uma conversa, mas com ritmo de suspense.
   - Use pausas naturais e frases curtas.
   - Mantenha o espectador curioso com perguntas retóricas.
   - Misture fatos reais e interpretações dramáticas, mas sem parecer teoria da conspiração.
   - Evite linguagem técnica; fale de forma acessível, como um narrador de documentário viral.

3. **Tom narrativo:**
   - Voz de mistério, levemente grave, com emoção e intensidade.
   - Mantenha um clima cinematográfico e envolvente, como um mini documentário.
   - Pode ser sombrio, inspirador ou místico dependendo do tema.
   - Transmita credibilidade, mesmo quando o conteúdo for misterioso.
   - Se o tema for sobre uma noticia neutra ou não muito interessante, invente algo para criar suspense e curiosidade.

4. **Conclusão / Chamada para ação:**
   - Termine com um fecho forte, que incentive o engajamento:
     - “Mande esse vídeo pra alguém que precisa saber disso.”
     - “Agora me diz… você acredita nisso?”
     - “Comente o que você faria.”
     - “Isso muda tudo, não é?”
   - Nunca finalize de forma neutra.

5. **Estilo técnico:**
   - Texto direto, sem marcações, bullets ou formatação.
   - Deve caber naturalmente em um vídeo de 45 a 60 segundos.
   - Ritmo rápido, mas sem perder clareza.
   - Pode conter uma leve atmosfera de suspense, terror, ou revelação espiritual, conforme o tema.
   - Escreva em tom afirmativo, cinematográfico e emocional.

⚙️ Instruções adicionais:
- Use descrições visuais e emocionais ("enquanto todos dormiam", "as luzes piscavam", "ninguém entendeu o que estava acontecendo").
- Se o tema for real, adicione detalhes que reforcem credibilidade (datas, locais, nomes).
- Se for fictício ou misterioso, trate como “possível verdade”, sem afirmar como farsa.
- Nunca use formatação Markdown, tópicos, números ou cabeçalhos.
- Entregue apenas o texto final do roteiro, como se fosse narrado em voz alta em um vídeo viral.

Retorne apenas o roteiro final pronto para narração, sem explicações adicionais, listas ou observações."""
            
            print(f"\n🤖 Gerando roteiro para: {tema}...")
            response = model.generate_content(prompt)
            roteiro = response.text.strip()
            
            print("✅ Roteiro gerado com sucesso!")
            return roteiro
        except Exception as e:
            print(f"❌ Erro ao gerar roteiro: {e}")
            return None

    def _atualizar_status_video(self, row_index):
        try:
            self.worksheet.cell(row=row_index, column=5, value="OK")
            self.worksheet.cell(row=row_index, column=7, value=datetime.now().strftime("%Y-%m-%d"))
            self.workbook.save(self.planilha_path)
            print(f"✅ Status atualizado na linha {row_index}.")
        except Exception as e:
            print(f"❌ Erro ao atualizar status: {e}")

    def processar_primeiro_tema(self):
        """Pega o primeiro tema da planilha que ainda não foi processado, gera o roteiro e atualiza o status."""
        if not self._abrir_planilha():
            return None

        if self.worksheet.max_row < 2:
            print("⚠️ Nenhum tema encontrado na planilha.")
            return None

        for row in range(2, self.worksheet.max_row + 1):
            roteiro_atual = self.worksheet.cell(row, 4).value
            video_pronto = self.worksheet.cell(row, 5).value

            if not roteiro_atual and not video_pronto:
                tema = self.worksheet.cell(row, 1).value
                descricao = self.worksheet.cell(row, 2).value

                if not tema:
                    print(f"⚠️ Tema vazio na linha {row}.")
                    continue

                print(f"\n📋 Processando tema da linha {row}:")
                print(f"   Tema: {tema}")
                print(f"   Descrição: {descricao or 'N/A'}")

                roteiro = self._gerar_roteiro(tema, descricao or "")

                if roteiro:
                    self.worksheet.cell(row, 4, roteiro)
                    self.workbook.save(self.planilha_path)
                    print(f"\n✅ Roteiro salvo na planilha na linha {row}!")
                    
                    # Video Pronto e Data serão atualizados apenas no Video_generator.py

                    print(f"\n--- Roteiro Gerado ---")
                    print(roteiro)
                    print("--- Fim do Roteiro ---")
                    
                    return roteiro.splitlines()[-1]
                else:
                    print("❌ Não foi possível gerar o roteiro.")
                    return None
        
        print("✅ Todos os temas já foram processados.")
        return None


def main():
    print("🎬 Video Creator - Gerando roteiro...")
    generator = RoteiroGenerator()
    resultado = generator.processar_primeiro_tema()
    if resultado:
        print(f"\n✅ Processo concluído! Última linha: '{resultado}'")
    else:
        print("\n⚠️ Não foi possível processar nenhum tema.")


if __name__ == "__main__":
    main()